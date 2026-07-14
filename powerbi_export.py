# ==================================================
# powerbi_export.py
# Excel Export + Power BI Integration for ProData AI
# Generates Power BI ready Excel workbooks
# ==================================================

import pandas as pd
import numpy as np
import io
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, LineChart, Reference
    from openpyxl.chart.series import DataPoint
    OPENPYXL_OK = True
except ImportError:
    OPENPYXL_OK = False


# ── Color constants ────────────────────────────────────────────────────────────
TEAL_DARK  = '00B894'
TEAL_LIGHT = 'E8FFF7'
DARK_BG    = '0F172A'
WHITE      = 'FFFFFF'
GRAY_LIGHT = 'F8FAFC'
GRAY_MID   = 'E2E8F0'
GRAY_DARK  = '475569'
INDIGO     = '6366F1'
AMBER      = 'F59E0B'
PINK       = 'EC4899'
GREEN      = '10B981'
RED        = 'EF4444'


def _header_fill(color=TEAL_DARK):
    return PatternFill(start_color=color, end_color=color, fill_type='solid')

def _light_fill(color=TEAL_LIGHT):
    return PatternFill(start_color=color, end_color=color, fill_type='solid')

def _white_fill():
    return PatternFill(start_color=WHITE, end_color=WHITE, fill_type='solid')

def _gray_fill():
    return PatternFill(start_color=GRAY_LIGHT, end_color=GRAY_LIGHT, fill_type='solid')

def _dark_fill():
    return PatternFill(start_color=DARK_BG, end_color=DARK_BG, fill_type='solid')

def _header_font(color=WHITE, size=11, bold=True):
    return Font(name='Calibri', bold=bold, size=size, color=color)

def _body_font(color='1E293B', size=10, bold=False):
    return Font(name='Calibri', bold=bold, size=size, color=color)

def _border():
    thin = Side(style='thin', color=GRAY_MID)
    return Border(left=thin, right=thin, top=thin, bottom=thin)

def _center():
    return Alignment(horizontal='center', vertical='center', wrap_text=True)

def _left():
    return Alignment(horizontal='left', vertical='center', wrap_text=True)

def _auto_width(ws, min_w=10, max_w=40):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_w, min(max_w, max_len + 2))

def _write_header_row(ws, row, headers, colors=None, start_col=1):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=start_col + i, value=h)
        color = colors[i] if colors and i < len(colors) else TEAL_DARK
        cell.fill = _header_fill(color)
        cell.font = _header_font()
        cell.alignment = _center()
        cell.border = _border()

def _write_data_row(ws, row, values, alt=False, start_col=1):
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=start_col + i, value=v)
        cell.fill = _gray_fill() if alt else _white_fill()
        cell.font = _body_font()
        cell.alignment = _left()
        cell.border = _border()


# ══════════════════════════════════════════════════════════════════════════════
# MAIN EXPORT FUNCTION
# ══════════════════════════════════════════════════════════════════════════════
def build_powerbi_excel(df, results, client_name='Client', project='Analysis'):
    """
    Builds a Power BI-ready Excel workbook with multiple sheets:
    1. README        - How to use this workbook in Power BI
    2. Raw_Data      - Cleaned dataset
    3. ML_Results    - Model leaderboard + metrics
    4. Feature_Importance - XAI driver scores
    5. Forecast_Data - Time series predictions
    6. Data_Profile  - Column statistics
    7. AI_Insights   - Claude AI analysis text
    8. Summary_KPIs  - Key metrics for Power BI card visuals
    """
    if not OPENPYXL_OK:
        raise Exception("openpyxl not installed. Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    now = datetime.now().strftime('%Y-%m-%d %H:%M')

    # ── Sheet 1: README ───────────────────────────────────────────────────────
    ws_read = wb.create_sheet('README')
    ws_read.sheet_properties.tabColor = TEAL_DARK

    # Title
    ws_read.merge_cells('A1:F1')
    c = ws_read['A1']
    c.value = 'ProData AI — Power BI Integration Workbook'
    c.fill = _dark_fill()
    c.font = Font(name='Calibri', bold=True, size=18, color=TEAL_DARK)
    c.alignment = _center()
    ws_read.row_dimensions[1].height = 40

    ws_read.merge_cells('A2:F2')
    c = ws_read['A2']
    c.value = f'Generated: {now}  |  Project: {project}  |  Client: {client_name}'
    c.fill = _header_fill(DARK_BG)
    c.font = _body_font(color='94A3B8', size=10)
    c.alignment = _center()

    instructions = [
        ('', ''),
        ('HOW TO USE THIS IN POWER BI', ''),
        ('Step 1', 'Open Power BI Desktop'),
        ('Step 2', 'Click Get Data → Excel Workbook'),
        ('Step 3', 'Select this file and click Open'),
        ('Step 4', 'In Navigator, select ALL sheets and click Load'),
        ('Step 5', 'Build visuals using the sheet data as described below'),
        ('', ''),
        ('SHEET GUIDE', ''),
        ('Raw_Data', 'Your cleaned dataset — use for all custom visuals and filters'),
        ('ML_Results', 'Model leaderboard — use for bar chart comparing model accuracy'),
        ('Feature_Importance', 'XAI drivers — use for horizontal bar chart showing what drives outcomes'),
        ('Forecast_Data', 'Predictions — use for line chart showing historical + forecast'),
        ('Data_Profile', 'Column statistics — use for table visual or card visuals'),
        ('AI_Insights', 'Claude AI analysis — use for text box or card in your report'),
        ('Summary_KPIs', 'Key numbers — connect directly to Power BI Card visuals'),
        ('', ''),
        ('RECOMMENDED POWER BI VISUALS', ''),
        ('Page 1: Overview', 'Cards from Summary_KPIs + Table from Raw_Data'),
        ('Page 2: ML Results', 'Bar chart from ML_Results (Model vs Score)'),
        ('Page 3: Key Drivers', 'Horizontal bar from Feature_Importance (Feature vs Importance)'),
        ('Page 4: Forecast', 'Line chart from Forecast_Data (Date vs Actual + Predicted)'),
        ('Page 5: AI Insights', 'Text card from AI_Insights sheet'),
    ]

    for i, (label, value) in enumerate(instructions):
        row = i + 3
        c1 = ws_read.cell(row=row, column=1, value=label)
        c2 = ws_read.cell(row=row, column=2, value=value)
        if label in ('HOW TO USE THIS IN POWER BI', 'SHEET GUIDE', 'RECOMMENDED POWER BI VISUALS'):
            for c in [c1, c2]:
                c.fill = _header_fill(INDIGO)
                c.font = _header_font()
            ws_read.merge_cells(f'A{row}:F{row}')
            c1.alignment = _left()
        elif label.startswith('Step') or label in ('Raw_Data','ML_Results','Feature_Importance','Forecast_Data','Data_Profile','AI_Insights','Summary_KPIs') or label.startswith('Page'):
            c1.font = _body_font(bold=True, color=TEAL_DARK)
            c2.font = _body_font()
            c1.fill = _gray_fill()
            c2.fill = _gray_fill()
        elif label == '':
            pass
        else:
            c1.font = _body_font(bold=True)
            c2.font = _body_font()

    ws_read.column_dimensions['A'].width = 28
    ws_read.column_dimensions['B'].width = 70

    # ── Sheet 2: Raw_Data ─────────────────────────────────────────────────────
    ws_data = wb.create_sheet('Raw_Data')
    ws_data.sheet_properties.tabColor = GREEN

    if df is not None and len(df) > 0:
        # Header
        headers = list(df.columns)
        _write_header_row(ws_data, 1, headers)
        ws_data.row_dimensions[1].height = 22

        # Data rows (limit to 50k for Excel performance)
        df_export = df.head(50000)
        for r_idx, row_data in enumerate(df_export.itertuples(index=False), start=2):
            alt = r_idx % 2 == 0
            for c_idx, val in enumerate(row_data, start=1):
                cell = ws_data.cell(row=r_idx, column=c_idx, value=val)
                cell.fill = _gray_fill() if alt else _white_fill()
                cell.font = _body_font()
                cell.alignment = _left()
                cell.border = _border()

        # Freeze header
        ws_data.freeze_panes = 'A2'
        _auto_width(ws_data)

    # ── Sheet 3: ML_Results ───────────────────────────────────────────────────
    ws_ml = wb.create_sheet('ML_Results')
    ws_ml.sheet_properties.tabColor = INDIGO

    ml = results.get('ml')
    if ml:
        # Title
        ws_ml.merge_cells('A1:G1')
        c = ws_ml['A1']
        c.value = f'AutoML Leaderboard — Target: {ml.get("target","")}'
        c.fill = _header_fill(INDIGO)
        c.font = _header_font(size=13)
        c.alignment = _center()
        ws_ml.row_dimensions[1].height = 30

        is_class = ml.get('is_class', False)
        if is_class:
            headers = ['Rank', 'Model', 'Accuracy', 'Status', 'Task Type']
        else:
            headers = ['Rank', 'Model', 'R² Score', 'MAE', 'RMSE', 'Status']

        _write_header_row(ws_ml, 2, headers)

        leaderboard = ml.get('leaderboard', [])
        medals = ['🥇 Winner', '🥈 2nd', '🥉 3rd']
        for i, row in enumerate(leaderboard):
            r = i + 3
            rank = medals[i] if i < 3 else f'#{i+1}'
            if is_class:
                vals = [rank, row.get('Model',''), row.get('Accuracy','-'),
                        row.get('Status',''), 'Classification']
            else:
                vals = [rank, row.get('Model',''), row.get('R²','-'),
                        row.get('MAE','-'), row.get('RMSE','-'), row.get('Status','')]
            _write_data_row(ws_ml, r, vals, alt=i%2==0)

            # Highlight winner
            if i == 0:
                for c_idx in range(1, len(vals)+1):
                    cell = ws_ml.cell(row=r, column=c_idx)
                    cell.fill = _light_fill(TEAL_LIGHT)
                    cell.font = _body_font(bold=True, color=TEAL_DARK)

        # Summary metrics
        skip = len(leaderboard) + 4
        ws_ml.cell(row=skip, column=1, value='WINNER SUMMARY').font = _header_font(color=TEAL_DARK)
        ws_ml.cell(row=skip+1, column=1, value='Best Model').font = _body_font(bold=True)
        ws_ml.cell(row=skip+1, column=2, value=ml.get('model_name','')).font = _body_font()
        if not is_class:
            ws_ml.cell(row=skip+2, column=1, value='R² Score').font = _body_font(bold=True)
            ws_ml.cell(row=skip+2, column=2, value=ml.get('r2','')).font = _body_font()
            ws_ml.cell(row=skip+3, column=1, value='MAE').font = _body_font(bold=True)
            ws_ml.cell(row=skip+3, column=2, value=ml.get('mae','')).font = _body_font()
            ws_ml.cell(row=skip+4, column=1, value='RMSE').font = _body_font(bold=True)
            ws_ml.cell(row=skip+4, column=2, value=ml.get('rmse','')).font = _body_font()
        else:
            ws_ml.cell(row=skip+2, column=1, value='Accuracy').font = _body_font(bold=True)
            ws_ml.cell(row=skip+2, column=2, value=ml.get('acc','')).font = _body_font()

        ws_ml.freeze_panes = 'A3'
        _auto_width(ws_ml)

    # ── Sheet 4: Feature_Importance ───────────────────────────────────────────
    ws_fi = wb.create_sheet('Feature_Importance')
    ws_fi.sheet_properties.tabColor = AMBER

    if ml and ml.get('importances'):
        ws_fi.merge_cells('A1:C1')
        c = ws_fi['A1']
        c.value = f'Business Driver Analysis — What drives {ml.get("target","your KPI")}?'
        c.fill = _header_fill(AMBER)
        c.font = _header_font(size=13)
        c.alignment = _center()
        ws_fi.row_dimensions[1].height = 30

        _write_header_row(ws_fi, 2, ['Rank', 'Feature / Driver', 'Importance Score'],
                         colors=[AMBER, AMBER, AMBER])

        importances = ml['importances']
        sorted_imp = sorted(importances.items(), key=lambda x: x[1], reverse=True)

        for i, (feat, score) in enumerate(sorted_imp):
            r = i + 3
            rank = f'#{i+1}'
            vals = [rank, feat, round(float(score), 6)]
            _write_data_row(ws_fi, r, vals, alt=i%2==0)
            # Highlight top driver
            if i == 0:
                for c_idx in range(1, 4):
                    cell = ws_fi.cell(row=r, column=c_idx)
                    cell.fill = PatternFill(start_color='FFF8E7', end_color='FFF8E7', fill_type='solid')
                    cell.font = _body_font(bold=True, color=AMBER)

        ws_fi.freeze_panes = 'A3'
        _auto_width(ws_fi)

    # ── Sheet 5: Forecast_Data ────────────────────────────────────────────────
    ws_fc = wb.create_sheet('Forecast_Data')
    ws_fc.sheet_properties.tabColor = PINK

    forecast = results.get('forecast')
    if forecast and forecast.get('fig'):
        ws_fc.merge_cells('A1:E1')
        c = ws_fc['A1']
        c.value = f'AI Forecast — {forecast.get("col","")}'
        c.fill = _header_fill(PINK)
        c.font = _header_font(size=13)
        c.alignment = _center()
        ws_fc.row_dimensions[1].height = 30

        _write_header_row(ws_fc, 2,
            ['Date', 'Type', 'Value', 'Lower Bound (95%)', 'Upper Bound (95%)'],
            colors=[PINK, PINK, PINK, PINK, PINK])

        # Extract data from plotly figure
        fig = forecast['fig']
        row_num = 3
        for trace in fig.data:
            trace_name = getattr(trace, 'name', '')
            x_vals = list(trace.x) if hasattr(trace, 'x') and trace.x is not None else []
            y_vals = list(trace.y) if hasattr(trace, 'y') and trace.y is not None else []

            if trace_name in ('Historical', 'Forecast') and x_vals and y_vals:
                for date, val in zip(x_vals, y_vals):
                    if val is not None:
                        _write_data_row(ws_fc, row_num,
                            [str(date)[:10], trace_name, round(float(val), 4), '', ''],
                            alt=row_num%2==0)
                        row_num += 1

        # Summary metrics
        info_row = row_num + 2
        ws_fc.cell(row=info_row, column=1, value='FORECAST SUMMARY').font = _header_font(color=PINK)
        ws_fc.cell(row=info_row+1, column=1, value='Metric').font = _body_font(bold=True)
        ws_fc.cell(row=info_row+1, column=2, value='Value').font = _body_font(bold=True)
        metrics = [
            ('Column', forecast.get('col', '')),
            ('Forecast Horizon', f"{forecast.get('horizon', 30)} days"),
            ('Latest Actual Value', round(float(forecast.get('latest', 0)), 2)),
            ('Projected Value', round(float(forecast.get('projected', 0)), 2)),
        ]
        for j, (k, v) in enumerate(metrics):
            ws_fc.cell(row=info_row+2+j, column=1, value=k).font = _body_font(bold=True)
            ws_fc.cell(row=info_row+2+j, column=2, value=v).font = _body_font()

        ws_fc.freeze_panes = 'A3'
        _auto_width(ws_fc)

    # ── Sheet 6: Data_Profile ─────────────────────────────────────────────────
    ws_prof = wb.create_sheet('Data_Profile')
    ws_prof.sheet_properties.tabColor = GRAY_DARK

    if df is not None:
        ws_prof.merge_cells('A1:H1')
        c = ws_prof['A1']
        c.value = 'Dataset Profile — Column Statistics'
        c.fill = _header_fill(GRAY_DARK)
        c.font = _header_font(size=13)
        c.alignment = _center()
        ws_prof.row_dimensions[1].height = 30

        _write_header_row(ws_prof, 2,
            ['Column', 'Type', 'Non-Null Count', 'Null Count', 'Null %',
             'Unique Values', 'Mean / Mode', 'Min / Max'])

        for i, col in enumerate(df.columns):
            r = i + 3
            dtype = str(df[col].dtype)
            non_null = int(df[col].notna().sum())
            null_count = int(df[col].isna().sum())
            null_pct = round(null_count / len(df) * 100, 1) if len(df) > 0 else 0
            unique = int(df[col].nunique())

            if df[col].dtype in [np.float64, np.int64, np.float32, np.int32]:
                mean_mode = round(float(df[col].mean()), 3) if df[col].notna().any() else ''
                min_max = f"{round(float(df[col].min()),2)} / {round(float(df[col].max()),2)}" if df[col].notna().any() else ''
            else:
                mode_vals = df[col].mode()
                mean_mode = str(mode_vals[0]) if len(mode_vals) > 0 else ''
                min_max = ''

            vals = [col, dtype, non_null, null_count, f'{null_pct}%', unique, mean_mode, min_max]
            _write_data_row(ws_prof, r, vals, alt=i%2==0)

            if null_pct > 20:
                ws_prof.cell(row=r, column=5).fill = PatternFill(start_color='FEE2E2', end_color='FEE2E2', fill_type='solid')
                ws_prof.cell(row=r, column=5).font = _body_font(color=RED, bold=True)

        ws_prof.freeze_panes = 'A3'
        _auto_width(ws_prof)

    # ── Sheet 7: AI_Insights ──────────────────────────────────────────────────
    ws_ai = wb.create_sheet('AI_Insights')
    ws_ai.sheet_properties.tabColor = INDIGO

    ws_ai.merge_cells('A1:B1')
    c = ws_ai['A1']
    c.value = 'AI-Generated Business Insights — ProData AI + Claude'
    c.fill = _header_fill(INDIGO)
    c.font = _header_font(size=13)
    c.alignment = _center()
    ws_ai.row_dimensions[1].height = 30

    insights_text = results.get('ai_insights', 'Run One-Click analysis with API key to generate AI insights.')
    lines = insights_text.split('\n') if insights_text else ['No insights generated yet.']

    for i, line in enumerate(lines):
        r = i + 2
        c = ws_ai.cell(row=r, column=1, value=line)
        c.font = _body_font()
        c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        ws_ai.row_dimensions[r].height = 20

    ws_ai.column_dimensions['A'].width = 120

    # ── Sheet 8: Summary_KPIs ─────────────────────────────────────────────────
    ws_kpi = wb.create_sheet('Summary_KPIs')
    ws_kpi.sheet_properties.tabColor = GREEN

    ws_kpi.merge_cells('A1:C1')
    c = ws_kpi['A1']
    c.value = 'Summary KPIs — Connect to Power BI Card Visuals'
    c.fill = _header_fill(GREEN)
    c.font = _header_font(size=13)
    c.alignment = _center()
    ws_kpi.row_dimensions[1].height = 30

    _write_header_row(ws_kpi, 2, ['KPI Name', 'Value', 'Category'],
                     colors=[GREEN, GREEN, GREEN])

    kpis = []

    if df is not None:
        kpis += [
            ('Total Rows', len(df), 'Dataset'),
            ('Total Columns', len(df.columns), 'Dataset'),
            ('Missing Values', int(df.isna().sum().sum()), 'Data Quality'),
            ('Missing %', round(df.isna().sum().sum() / (len(df)*len(df.columns)) * 100, 1), 'Data Quality'),
            ('Numeric Columns', len(df.select_dtypes(include=np.number).columns), 'Dataset'),
            ('Categorical Columns', len(df.select_dtypes(exclude=np.number).columns), 'Dataset'),
        ]

    if ml:
        is_class = ml.get('is_class', False)
        kpis += [
            ('Best ML Model', ml.get('model_name', ''), 'ML Results'),
            ('Models Tested', ml.get('n_models', 6), 'ML Results'),
            ('Target Variable', ml.get('target', ''), 'ML Results'),
            ('Task Type', 'Classification' if is_class else 'Regression', 'ML Results'),
        ]
        if not is_class and ml.get('r2') not in (None, '-'):
            kpis += [
                ('R² Score', ml.get('r2', ''), 'ML Performance'),
                ('MAE', ml.get('mae', ''), 'ML Performance'),
                ('RMSE', ml.get('rmse', ''), 'ML Performance'),
            ]
        elif is_class and ml.get('acc') not in (None, '-'):
            kpis += [('Accuracy', ml.get('acc', ''), 'ML Performance')]

    if ml and ml.get('importances'):
        top_driver = list(ml['importances'].keys())[0] if ml['importances'] else ''
        kpis += [('Top Business Driver', top_driver, 'XAI')]

    if forecast:
        kpis += [
            ('Forecast Column', forecast.get('col', ''), 'Forecast'),
            ('Latest Value', round(float(forecast.get('latest', 0)), 2), 'Forecast'),
            ('Projected Value', round(float(forecast.get('projected', 0)), 2), 'Forecast'),
            ('Forecast Horizon (days)', forecast.get('horizon', 30), 'Forecast'),
        ]

    kpis += [
        ('Generated By', 'ProData AI', 'Meta'),
        ('Generated At', datetime.now().strftime('%Y-%m-%d %H:%M'), 'Meta'),
        ('Project', project, 'Meta'),
        ('Client', client_name, 'Meta'),
    ]

    for i, (name, val, cat) in enumerate(kpis):
        r = i + 3
        _write_data_row(ws_kpi, r, [name, val, cat], alt=i%2==0)
        ws_kpi.cell(row=r, column=1).font = _body_font(bold=True)

    ws_kpi.freeze_panes = 'A3'
    ws_kpi.column_dimensions['A'].width = 30
    ws_kpi.column_dimensions['B'].width = 30
    ws_kpi.column_dimensions['C'].width = 20

    # ── Save to bytes ─────────────────────────────────────────────────────────
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()
